from openpyxl import workbook, load_workbook
import time
import os
import pyautogui
import webbrowser
import logging
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
    )

from userQuestions import *
from load import *
from Logos import *
from databaseBACKUP import *

logging.basicConfig(level=logging.DEBUG,  
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename='backup+log/BoxNest.log', filemode='w')

word = "database"

workbook = load_workbook(f"{word}.xlsx")
sheet = workbook.active

logging.info("BoxNest: Script started")

logging.info("Initiating database loading sequence")
clear_cmd()
logging.info("Clearing command prompt for a clean interface")
load_sequence()
logging.info("Loading sequence displayed")
clear_cmd()
logging.info("Clearing command prompt for a clean interface")
Logo()
logging.info("Displaying logo for brand recognition")

logging.info("Loading sequence completed successfully")


while True:
    logging.info("Entering main loop for user interaction")
    CheckBackup()
    last_row = sheet.max_row
    box = sheet.max_row - 1

    Choice_Box()
    logging.info("User interface: Choice_Box displayed")

    
    Choice = questionChoicesLoop("Please Select An Option:", "add", ["add", "find", "remove", "open", "quit", "website", "print", "1", "2", "3", "4", "5", "6", "7"]).lower().strip()
    
    logging.info(f"User made a choice. Selected option: {Choice}")

    if Choice == "add" or Choice == "1":
        print(f"The last added was box number: {box}")
        box = box + 1
        box = str(box)
        BoxNum = questionIntLoop("What is the number of this box being added:", box)
        Building = questionChoicesLoop("What building is it in (main, left, right):", "main", ["main", "left", "right"])
        Row = questionIntWithRangeLoop("which row is the item located in   (0 - 9):", "1", 0, 9)
        level = questionIntWithRangeLoop("which level is the item located on (1 - 5):", "1", 0, 5)
        Size = questionIntWithRangeLoop("What is the size of the items box  (0 - 9):", "1", 0, 9)
        Owner = questionStringLoop("Who is the owner of the box being added   :", "NAME")
        description = questionStringBlank("To not put a description press 'enter' or\nDescribe what's inside the box being added:", "")

        logging.info("User interaction: Questions asked")

        if Building == "main": 
            Building_num = 1
        elif Building == "left":
            Building_num = 2 
        elif Building == "right":
            Building_num = 3
        else:
            Building_num = 0

        BoxID = f"{BoxNum}{Building_num}{Row}{level}{Size}"
        BoxID = int(BoxID)

        print(f"The ID of the box is {BoxID}", sep='')
        

        line = BoxNum + 1

        sheet[f'A{line}'] = BoxID
        sheet[f'B{line}'] = Building
        sheet[f'C{line}'] = Row
        sheet[f'D{line}'] = level
        sheet[f'E{line}'] = Size
        sheet[f'F{line}'] = Owner

        if description == "":
            sheet[f'G{line}'] = description
        else:
            sheet[f'G{line}'] = f"Box of {description}"

        workbook.save(f"{word}.xlsx")

        print(f"BoxID {BoxID} created and added to the database")

        logging.info(f"Data processing: BoxID {BoxID} created and added to the database")

        border = rf"""
+------------------------------------+
| ____            _   _          _   |
|| __ )  _____  _| \ | | ___ ___| |_ |
||  _ \ / _ \ \/ |  \| |/ _ / __| __||
|| |_) | (_) >  <| |\  |  __\__ | |_ |
||____/ \___/_/\_|_| \_|\___|___/\__||
|                                    |
|           Box ID: {BoxID}           |
+------------------------------------+
"""

        label = open("label.txt", "w")
        label.write(border)
        label.close()

        logging.info("Document generation: label.txt created")


#        os.startfile("label.txt")
        os.startfile("label.txt", "print")
        
        print("Printing box ID sticker...")

        logging.info("Document printing: Box ID sticker printed")

        time.sleep(0.25)

        openDoc = str(questionString("Type 'open' to open the database or press enter to continue:","continue"))

        if openDoc == "open":
            logging.info("Database access: Database opened")
            time.sleep(0.25)
            print("Opening file...")
            os.startfile(f"{word}.xlsx")
            
            while True:
                time.sleep(1)
                close = str(questionString("Press enter to close the sheet:","close"))
                if close == "close":
                    logging.info("Database access: Database closed")
                    print("closing sheet...")
                    pyautogui.hotkey('alt', 'tab')
                    pyautogui.hotkey('ctrl', 's')
                    time.sleep(0.25)
                    pyautogui.hotkey('alt', 'f4')
                    clear_cmd()
                    break 
        elif openDoc == "continue":
            logging.info("Database access: Database not opened; program continues")
            print("Continuing...")
            clear_cmd()
        else:
            logging.info("Database access: Database not opened; program continues")
            print("Not opening file...")
            clear_cmd()
                
    elif Choice == "find" or Choice == "2":
        print("To find a box you need to know the ID")
        find = questionChoicesLoop("Do you know the ID of the box:", "Yes", ["yes", "no"]).lower().strip()
        find = find.upper()

        if find == "YES":
            logging.info("User interaction: User knows ID")
            ID = questionIntLoop("What is the ID number of the box:","")

            print(f"Searching for a box with ID: {ID}...")
            logging.info(f"User interaction: Searching for a box with ID: {ID}")
            time.sleep(0.50)
            for row in range(2, last_row + 1):
                cell = sheet.cell(row, 1)
                if cell.value == ID:
                    st = [1]
                    nd = [2]
                    rd = [3]
                    th = [4, 5, 6, 7, 8 ,9]

                    if sheet[f'C{row}'].value in st:
                        end = "st"
                    elif sheet[f'C{row}'].value in nd:
                        end = "nd"
                    elif sheet[f'C{row}'].value in rd:
                        end = "rd"
                    elif sheet[f'C{row}'].value in th:
                        end = "th"
                    
                    if sheet[f'D{row}'].value in st:
                        end = "st"
                    elif sheet[f'D{row}'].value in nd:
                        end = "nd"
                    elif sheet[f'D{row}'].value in rd:
                        end = "rd"
                    elif sheet[f'D{row}'].value in th:
                        end = "th"

                    logging.info(f"Box {ID} found in row {row} of the database")

                    print(f"Box {ID} is in row {row} of the database")
                    time.sleep(1)
                    print(f"It is in the {sheet[f'B{row}'].value} building on the {sheet[f'C{row}'].value}{end} row on the {sheet[f'D{row}'].value}{end} level and it is a box size {sheet[f'E{row}'].value}")
                    time.sleep(3)
                    clear_cmd()
                    break
            else:
                print(f"A box with the ID: {ID} has not found")
                logging.info(f"Box {ID} not found in database")
                time.sleep(1)
                clear_cmd()
        elif find == "NO":
            print("You need to know the ID of the box to find it")
            logging.info("User interaction: User does not know the ID of the box")
            time.sleep(0.25)
            clear_cmd()

    elif Choice == "remove" or Choice == "3":
        ID = questionIntLoop("What is the ID number of the box:","")
        
        print(f"Searching for a box with ID: {ID}...")
        logging.info(f"User interaction: Searching for a box with ID: {ID}")
        time.sleep(0.50)
        for row in range(2, last_row + 1):
            cell = sheet.cell(row, 1)
            if cell.value == ID:
                st = [1]
                nd = [2]
                rd = [3]
                th = [4, 5, 6, 7, 8 ,9]

                if sheet[f'C{row}'].value in st:
                    end = "st"
                elif sheet[f'C{row}'].value in nd:
                    end = "nd"
                elif sheet[f'C{row}'].value in rd:
                    end = "rd"
                elif sheet[f'C{row}'].value in th:
                    end = "th"
                
                if sheet[f'D{row}'].value in st:
                    end = "st"
                elif sheet[f'D{row}'].value in nd:
                    end = "nd"
                elif sheet[f'D{row}'].value in rd:
                    end = "rd"
                elif sheet[f'D{row}'].value in th:
                    end = "th"

                print(f"Box {ID} is in row {row} of the database")
                logging.info(f"Box {ID} found in row {row} of the database")
                time.sleep(1)
                print(f"It is in the {sheet[f'B{row}'].value} building on the {sheet[f'C{row}'].value}{end} row on the {sheet[f'D{row}'].value}{end} level and it is a box size {sheet[f'E{row}'].value}")
                time.sleep(1)

                remove = questionChoicesLoop("Are you sure you want to remove this box:", "yes", ["yes", "no"]).lower().strip()
                if remove == "yes":
                    sheet.delete_rows(row, 1)
                    workbook.save(f"{word}.xlsx")
                    logging.info(f"Database access: Box {ID} removed from database")
                    print("Box removed")
                    time.sleep(0.25)
                    clear_cmd()
                    break
                elif remove == "no":
                    logging.info(f"Database access: Box {ID} not removed from database")
                    print("Box not removed")
                    time.sleep(0.25)
                    clear_cmd()
                    break
        else:
            print(f"A box with the ID: {ID} has not been found")
            logging.info(f"Database access: Box {ID} not found in database")
            time.sleep(1)
            clear_cmd()

    elif Choice == "open" or Choice == "4": 
        time.sleep(0.25)
        print("Opening file...")
        os.startfile(f"{word}.xlsx")
        logging.info("Database access: Database opened")
        
        while True:
            time.sleep(1)
            close = str(questionString("Press enter to close the sheet:","close"))
            if close == "close":
                logging.info("Database access: Database closed")
                print("closing sheet...")
                pyautogui.hotkey('alt', 'tab')
                pyautogui.hotkey('ctrl', 's')
                time.sleep(0.25)
                pyautogui.hotkey('alt', 'f4')
                clear_cmd()
                break 
                
    elif Choice == "quit" or Choice == "5":
        logging.critical("BoxNest: script has successfully finished")

        def type_text(text, delay=0.01):
            for char in text:
                print(char, end='', flush=True)
                time.sleep(delay)
            print()

        Quitting = "Quitting program..."
        Thanks = "Thank you for using BoxNest"
        Creator = "Creator: Kyle Seaford"
        type_text(Thanks)
        type_text(Creator)
        type_text(Quitting)
        time.sleep(0.25)
        break

    elif Choice == "website" or Choice == "6":
        logging.info("BoxNest: website opened successfully")
        print("Opening website...")
        webbrowser.open('http://test.kyle-seaford.co.uk/index.html')
        time.sleep(1)
        clear_cmd()

    elif Choice == "print" or Choice == "7":
        print("printing database...")
        logging.info("BoxNest: database printed successfully")
        os.startfile("database.xlsx")
#        os.startfile("database.xlsx", "print")
        clear_cmd()